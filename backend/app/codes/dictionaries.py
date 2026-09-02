from __future__ import annotations

import re
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

import openpyxl


def normalize_icd10(code: str) -> str:
    return code.strip().upper().replace(".", "")


@dataclass(frozen=True)
class CodeDictionaries:
    cpt: frozenset[str]
    hcpcs: frozenset[str]
    icd10: frozenset[str]
    descriptions: dict[str, str]

    def contains(self, code: str) -> str | None:
        raw = code.strip().upper()
        if raw in self.cpt:
            return "cpt"
        if raw in self.hcpcs:
            return "hcpcs"
        if normalize_icd10(raw) in self.icd10:
            return "icd10"
        return None


def _load_cpt_xlsx(path: Path) -> tuple[set[str], dict[str, str]]:
    codes: set[str] = set()
    descriptions: dict[str, str] = {}
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        code = str(row[0]).strip().upper()
        if not code:
            continue
        codes.add(code)
        if len(row) > 1 and row[1]:
            descriptions.setdefault(code, str(row[1]).strip())
    workbook.close()
    return codes, descriptions


def _load_hcpcs_txt(path: Path) -> tuple[set[str], dict[str, str]]:
    codes: set[str] = set()
    descriptions: dict[str, str] = {}
    with path.open(encoding="latin-1") as fh:
        for line in fh:
            stripped = line.strip()
            if len(stripped) < 5:
                continue
            code = stripped[:5].upper()
            if not re.fullmatch(r"[A-Z0-9]{5}", code):
                continue
            codes.add(code)
            desc = stripped[8:].strip()
            if desc:
                descriptions.setdefault(code, desc[:120].strip())
    return codes, descriptions


def _load_icd10_order(path: Path) -> tuple[set[str], dict[str, str]]:
    codes: set[str] = set()
    descriptions: dict[str, str] = {}
    with path.open(encoding="latin-1") as fh:
        for line in fh:
            if len(line) < 20:
                continue
            code = line[6:13].strip().upper()
            if not re.fullmatch(r"[A-Z][0-9A-Z]{2,6}", code):
                continue
            codes.add(code)
            long_desc = line[77:].strip() or line[16:77].strip()
            if long_desc:
                descriptions.setdefault(code, long_desc)
    return codes, descriptions


def load_dictionaries(reference_root: Path) -> CodeDictionaries:
    reference_root = Path(reference_root)
    cpt, cpt_desc = _load_cpt_xlsx(reference_root / "cpt-codes" / "cpt.xlsx")
    hcpcs, hcpcs_desc = _load_hcpcs_txt(reference_root / "cpt-codes" / "cpt-codes.txt")
    icd, icd_desc = _load_icd10_order(reference_root / "ict-10-codes" / "icd10cm_order_2026.txt")
    descriptions = {**hcpcs_desc, **cpt_desc, **icd_desc}
    return CodeDictionaries(frozenset(cpt), frozenset(hcpcs), frozenset(icd), descriptions)


@lru_cache
def get_dictionaries() -> CodeDictionaries:
    from ..config import get_settings

    return load_dictionaries(get_settings().reference_root)
